import os
import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from dateutil.parser import parse

# File Paths
PRE_UPDATED = "C:\\Users\\sivaps15\\OneDrive - McMaster University\\Billing\\Pre-Updated"
INTERMEDIATE_FOLDER = "C:\\Users\\sivaps15\\OneDrive - McMaster University\\Billing\\Intermediate Folder"
OUTPUT_FOLDER = "C:\\Users\\sivaps15\\OneDrive - McMaster University\\Billing\\Output"

# Check that required folders exist
for path in [PRE_UPDATED, INTERMEDIATE_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(path):
        # raise FileNotFoundError(f"❌ Please create folder: {path}")
                raise FileNotFoundError(
            f"\n❌ Please create folder '{os.path.basename(path)}' at:\n{path}\n"
        )
os.makedirs(INTERMEDIATE_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def extract_clean_meter_name(raw_name):
    parts = raw_name.split("M")
    if len(parts) > 1 and parts[1][:1].isdigit():
        return "M".join([parts[0], parts[1][:1]]) + " " + parts[1][1:].strip()
    return raw_name

def clean_building_name(filename):
    base = os.path.splitext(filename)[0]
    name_part = base.split("Residence - ")[-1]
    name_part = re.sub(r"\b(?:Report|report|timestamp)\b", "", name_part, flags=re.IGNORECASE)
    name_part = re.sub(r"\b\d{4}[- ]\d{2}[- ]\d{2}\b", "", name_part)
    name_part = re.sub(r"\b\d{1,2}[- ]\d{1,2}[- ]\d{2,4}\b", "", name_part)
    name_part = re.sub(r"\b\d{4}\b", "", name_part)
    name_part = re.sub(r"\b\d{1,2}\b", "", name_part)
    name_part = re.sub(r"[_\-]{2,}", " ", name_part)
    name_part = re.sub(r"\s{2,}", " ", name_part)
    return name_part.strip(" -_")
def round_to_nearest_power_of_10(val, is_cogen):
    power = 10 ** (len(str(int(abs(val)))) - 1)
    if is_cogen:
        return math.floor(val / power) * power
    else:
        return math.ceil(val / power) * power

def format_excel(input_path, intermediate_subfolder, master_data, building_name, today_str, bill_month, time_str):
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from dateutil.parser import parse
    import os

    filename = os.path.basename(input_path)
    print(f"🔄 Processing file: {filename}")
    temp_path = input_path

    if input_path.lower().endswith(".xls"):
        df = pd.read_excel(input_path, engine="xlrd")
        temp_path = os.path.join(intermediate_subfolder, filename.replace(".xls", "_temp.xlsx"))
        df.to_excel(temp_path, index=False)

    wb = load_workbook(temp_path)
    sheet = wb.active

    timestamp_row = None
    for row in sheet.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and "timestamp" in str(cell.value).lower():
                timestamp_row = cell.row
                break
        if timestamp_row:
            break

    if timestamp_row is None:
        print(f"⚠️ Timestamp not found in: {filename}")
        return

    irya_start = 1
    for i in range(1, timestamp_row):
        if sheet.cell(row=i, column=1).value and "Information Requiring Your Attention" in str(sheet.cell(row=i, column=1).value):
            irya_start = i
            sheet.cell(row=i, column=1).font = Font(bold=True)

    if irya_start > 1:
        for _ in range(irya_start - 1):
            sheet.delete_rows(1)
        timestamp_row -= (irya_start - 1)

    first_data_row = timestamp_row + 1
    last_data_row = sheet.max_row

    for col_idx in reversed(range(2, sheet.max_column + 1)):
        col_values = [sheet.cell(row=r, column=col_idx).value for r in range(first_data_row, last_data_row + 1)]
        if all((v is None or str(v).strip() == "") for v in col_values):
            sheet.delete_cols(col_idx)

    meter_labels = [cell.value for cell in sheet[timestamp_row]]
    usage_row_index = timestamp_row - 1

    usage_cells = []

    for i, col in enumerate(sheet.iter_cols(min_row=first_data_row, min_col=2), start=2):
        values = [(row, cell.value) for row, cell in enumerate(col, start=first_data_row) if isinstance(cell.value, (int, float))]
        if len(values) < 2:
            continue

        first_row, first = values[0]
        last_row, last = values[-1]

        # Flip detection
        flip_value = None
        for j in range(1, len(values)):
            prev_val = values[j - 1][1]
            curr_val = values[j][1]
            if prev_val != 0 and abs(curr_val / prev_val) < 0.1:
                flip_value = prev_val
                break

        is_cogen = "cogen" in building_name.lower()
        if flip_value and ((last < first and not is_cogen) or (last > first and is_cogen)):
            rounded = round_to_nearest_power_of_10(flip_value, is_cogen)
            usage = (rounded - first) + last
        else:
            usage = last - first

        raw_meter_name = meter_labels[i - 1] if i - 1 < len(meter_labels) else f"Meter {i}"
        clean_meter = extract_clean_meter_name(str(raw_meter_name))

        try:
            first_time = parse(str(sheet.cell(row=first_row, column=1).value))
            last_time = parse(str(sheet.cell(row=last_row, column=1).value))
            next_month = 1 if first_time.month == 12 else first_time.month + 1
            next_year = first_time.year + 1 if first_time.month == 12 else first_time.year
            correct = (
                first_time.day == 1 and first_time.strftime("%I:%M %p") == "12:15 AM" and
                last_time.day == 1 and last_time.month == next_month and last_time.year == next_year and
                last_time.strftime("%I:%M %p") == "12:00 AM"
            )
        except:
            correct = False

        color = "C6EFCE" if correct else "FFC7CE"
        usage_cell = sheet.cell(row=usage_row_index, column=col[0].column)
        usage_cell.value = round(usage, 2)
        usage_cell.number_format = '#,##0.00'
        usage_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        usage_cells.append(usage_cell)

        master_data.append([building_name, clean_meter, round(usage, 2)])

    # Add table
    end_col = get_column_letter(sheet.max_column)
    end_row = sheet.max_row
    table = Table(displayName="MeterTable", ref=f"A{timestamp_row}:{end_col}{end_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Force RIGHT alignment on usage cells AFTER table is created
    for cell in usage_cells:
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Column width and formatting
    for col in sheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if cell.row != usage_row_index:
                cell.alignment = Alignment(wrap_text=True)
        sheet.column_dimensions[col_letter].width = max_len + 2

    output_filename = f"{today_str}_{time_str}_{bill_month}_{building_name}.xlsx"
    output_path = os.path.join(intermediate_subfolder, output_filename)
    wb.save(output_path)

    if temp_path != input_path:
        os.remove(temp_path)

    print(f"✅ Completed file: {filename}")


def main():
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H-%M")
    bill_month = now.strftime("%B")
    master_data = []

    intermediate_subfolder_name = f"{today_str}_{time_str}_{bill_month}"
    intermediate_subfolder = os.path.join(INTERMEDIATE_FOLDER, intermediate_subfolder_name)
    os.makedirs(intermediate_subfolder, exist_ok=True)

    files = [file for file in os.listdir(PRE_UPDATED) if file.endswith((".xls", ".xlsx"))]
    print(f"📁 Found {len(files)} files in Pre-Updated folder.")
    for idx, file in enumerate(files, start=1):
        building = clean_building_name(file)
        file_path = os.path.join(PRE_UPDATED, file)
        format_excel(file_path, intermediate_subfolder, master_data, building, today_str, bill_month, time_str)
        os.remove(file_path)
        print(f"📦 {idx}/{len(files)} files processed.")

    if master_data:
        master_df = pd.DataFrame(master_data, columns=["Building", "Meter", "Usage"])
        master_df["Usage"] = master_df["Usage"].map(lambda x: f"{x:,.2f}")
        master_filename = f"Final-{today_str}-{time_str}-{bill_month}.xlsx"
        master_path = os.path.join(OUTPUT_FOLDER, master_filename)

        with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
            master_df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[col_letter].width = max_length + 2

        print(f"✅ Master Excel file saved: {master_filename}")

if __name__ == "__main__":
    main()
