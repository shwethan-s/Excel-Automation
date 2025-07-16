def handle_nitrogen_file(input_path, intermediate_subfolder, building, today_str, bill_month, time_str):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import os
    from uuid import uuid4

    filename = os.path.basename(input_path)
    base_name = os.path.splitext(filename)[0]
    output_filename = f"{base_name}_cleaned.xlsx"
    output_path = os.path.join(intermediate_subfolder, output_filename)

    df = pd.read_excel(input_path, header=None)
    title = str(df.iloc[0, 0]).strip()
    headers = df.iloc[1].fillna("").astype(str).tolist()

    seen = {}
    clean_headers = []
    for i, h in enumerate(headers):
        h = h.strip()
        if not h or h.lower() == "nan":
            h = f"Meter{i+1}"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 1
        clean_headers.append(h)

    data = df.iloc[2:].copy()
    data.columns = clean_headers
    totals_row = data.iloc[-1]
    data = data.iloc[:-1]
    data.dropna(axis=1, how='all', inplace=True)
    clean_headers = list(data.columns)

    temp_path = os.path.join(intermediate_subfolder, f"temp_{uuid4().hex[:6]}.xlsx")
    data.to_excel(temp_path, index=False)

    wb = load_workbook(temp_path)
    ws = wb.active

    ws.insert_rows(1)
    ws.cell(row=1, column=1).value = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    usage_values = []
    for col in clean_headers:
        if col.lower() == "timestamp":
            usage_values.append("Usage")
        else:
            try:
                val = float(totals_row[col])
                usage_values.append(round(val, 2))
            except:
                usage_values.append("")

    ws.insert_rows(2)
    for col_idx, val in enumerate(usage_values, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        if isinstance(val, (int, float)):
            cell.number_format = "#,##0.00"

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col in enumerate(ws.iter_cols(min_row=4, min_col=1), start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(cell.value, str) and col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)

    try:
        start_row = 3
        end_row = ws.max_row
        last_col = ws.max_column
        table_range = f"A{start_row}:{get_column_letter(last_col)}{end_row}"
        table = Table(displayName=f"Table_{uuid4().hex[:6]}", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    except Exception as e:
        print(f"⚠️ Table formatting failed: {e}")

    wb.save(output_path)
    os.remove(temp_path)
    print(f"✅ Cleaned nitrogen file saved: {base_name}_cleaned.xlsx")
    return output_path
