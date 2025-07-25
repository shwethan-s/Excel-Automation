import os
import re
import math
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from dateutil.parser import parse
from uuid import uuid4


# Get OneDrive root path
onedrive_root = os.environ.get("OneDrive")
if not onedrive_root:
    raise EnvironmentError("❌ OneDrive path not found. Please ensure OneDrive is set up on this user account.")

# Log folder path using OneDrive
LOG_FOLDER = os.path.join(onedrive_root, "Excel Automation Tool", "Logs")
os.makedirs(LOG_FOLDER, exist_ok=True)
master_log_path = os.path.join(LOG_FOLDER, "master_log.txt")




class EmojiFilter(logging.Filter):
    def filter(self, record):
       
        record.msg = re.sub(r'[^\x00-\x7F]+', '', str(record.msg))


        return True

formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s')

# File handler - this will log to a file and strip emojis
file_handler = logging.FileHandler(master_log_path, mode='a', encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)
file_handler.addFilter(EmojiFilter())

# Console handler - this will log to the console and strip emojis
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)


logging.basicConfig(level=logging.INFO, handlers=[file_handler])


# File Paths
onedrive_root = os.environ.get("OneDrive")
if not onedrive_root:
    raise EnvironmentError("❌ OneDrive path not found. Please ensure OneDrive is set up on this user account.")

# Defined folder paths for folders 
PRE_UPDATED = os.path.join(onedrive_root, "Excel Automation Tool", "Pre-Updated")
INTERMEDIATE_FOLDER = os.path.join(onedrive_root, "Excel Automation Tool", "Intermediate Folder")
OUTPUT_FOLDER = os.path.join(onedrive_root, "Excel Automation Tool", "Output")

# Check that required folders exist
for path in [PRE_UPDATED, INTERMEDIATE_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"\n❌ Please create folder '{os.path.basename(path)}' at:\n{path}\n")
os.makedirs(INTERMEDIATE_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def extract_clean_meter_name(raw_name):
    parts = raw_name.split("M")
    if len(parts) > 1 and parts[1][:1].isdigit():
        return "M".join([parts[0], parts[1][:1]]) + " " + parts[1][1:].strip()
    return raw_name


def clean_building_name(filename):
    base = os.path.splitext(filename)[0]

    if re.match(r"^nitrogen\s+\d+[a-z]?$", base.lower()):

        return base.strip()
        

    name_part = base.split("Residence - ")[-1]
    name_part = re.sub(r"\b(?:Report|report|timestamp)\b", "", name_part, flags=re.IGNORECASE)
    name_part = re.sub(r"\b\d{4}[- ]\d{2}[- ]\d{2}\b", "", name_part)
    name_part = re.sub(r"\b\d{1,2}[- ]\d{1,2}[- ]\d{2,4}\b", "", name_part)
    name_part = re.sub(r"\b\d{4}\b", "", name_part)
    name_part = re.sub(r"\b\d{1,2}\b", "", name_part)
    name_part = re.sub(r"[_\-]{2,}", " ", name_part)
    name_part = re.sub(r"\s{2,}", " ", name_part)
    return name_part.strip(" -_")

def round_to_nearest_power_of_10(val, is_cogen):
    power = 10 ** (len(str(int(abs(val)))) - 1)
    return math.floor(val / power) * power if is_cogen else math.ceil(val / power) * power

def handle_nitrogen_file(input_path, intermediate_subfolder, building, today_str, bill_month, time_str, master_data):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import os
    from uuid import uuid4
    import logging
    from dateutil.parser import parse

    filename = os.path.basename(input_path)
    print(f"🔄 Processing file: {filename}")
    logging.info(f"Processing file: {filename}")

    # Load full sheet
    df = pd.read_excel(input_path, header=None)

    # Extract Nitrogen title
    title_row_values = df.iloc[0].dropna().astype(str).tolist()
    title_text = " – ".join(title_row_values).strip()
    df = df[1:]

    # Find timestamp row
    timestamp_row_idx = None
    for idx, row in df.iterrows():
        if any(isinstance(val, str) and "timestamp" in str(val).lower() for val in row):
            timestamp_row_idx = idx
            break
    if timestamp_row_idx is None:
        print(f"⚠️ Timestamp not found in: {filename}")
        logging.warning(f"Timestamp not found in: {filename}")
        return

    # Extract and format data from timestamp row
    df = df.iloc[timestamp_row_idx:].reset_index(drop=True)
    headers = df.iloc[0].fillna("").astype(str).tolist()
    df = df[1:].reset_index(drop=True)
    df.columns = headers
    df = df.dropna(axis=1, how='all')
    df = df.rename(columns={df.columns[0]: "Timestamp"})

    # Extract bill_month from first valid timestamp
    extracted_month = bill_month  # fallback
    for val in df["Timestamp"]:
        try:
            dt = parse(str(val), fuzzy=True)
            extracted_month = dt.strftime("%B")
            break
        except:
            continue

    # Compute usage
    usage_row = {"Timestamp": "Usage"}



    for col in df.columns[1:]:
        col_data = pd.to_numeric(df[col], errors="coerce").dropna()
        usage_row[col] = round(col_data.iloc[-1], 2) if not col_data.empty else ""
    
    # Append usage values to master_data (for Final output file)
    if master_data is not None:
        for col in df.columns[1:]:
            col_data = pd.to_numeric(df[col], errors="coerce").dropna()
            if not col_data.empty:
                usage = round(col_data.iloc[-1], 2)
                meter_name = str(col).strip()
                master_data.append([building, meter_name, usage])


    # Combine final data
    final_df = pd.concat([pd.DataFrame([usage_row]), df], ignore_index=True)

    # Use same naming format as other files
    output_filename = f"{today_str}_{time_str}_{extracted_month}_{building}.xlsx"
    output_path = os.path.join(intermediate_subfolder, output_filename)

    temp_path = os.path.join(intermediate_subfolder, f"{building.replace(' ', '_')}_temp.xlsx")
    final_df.to_excel(temp_path, index=False)

    # Format with openpyxl
    wb = load_workbook(temp_path)
    sheet = wb.active
    sheet.insert_rows(1)
    sheet.cell(row=1, column=1).value = title_text
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for col in range(2, sheet.max_column + 1):
        cell = sheet.cell(row=2, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx, cell in enumerate(sheet[3], start=1):
        val = str(cell.value).strip()
        cell.value = val if val else f"Column{col_idx}"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            sheet.column_dimensions[col_letter].width = 28
        else:
            sheet.column_dimensions[col_letter].width = 35



    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                #cell.alignment = Alignment (horizontal = "middle". uppercase(), vertical="center")
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.column == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[3]:
        cell.value = str(cell.value)

    for col_idx in range(2, sheet.max_column + 1):
        cell = sheet.cell(row=3, column=col_idx)
        # ensure it’s a float
        try:
            cell.value = float(cell.value)
        except:
            cell.value = 0.0
        # apply comma separators & two decimals
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right', vertical='center')


    if sheet.max_row >= 4:
        try:
            last_col = get_column_letter(sheet.max_column)
            last_row = sheet.max_row
            # START AT ROW 2 (real headers), not row 3
            table_range = f"A2:{last_col}{last_row}"
            table = Table(
                displayName=f"MeterTable_{uuid4().hex[:6]}",
                ref=table_range
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            sheet.add_table(table)
        except Exception as e:
            print(f"⚠️ Could not apply table style to {filename}: {e}")
            logging.warning(f"Could not apply table style to {filename}: {e}")

    wb.save(output_path)
    os.remove(temp_path)


    print(f"✅ Completed file: {filename}")
    logging.info(f"Completed file: {filename} - saved to {output_path}")


def format_excel(input_path, intermediate_subfolder, master_data, building_name, today_str, bill_month, time_str):
    filename = os.path.basename(input_path)
    print(f"🔄 Processing file: {filename}")
    logging.info(f"Processing file: {filename}")

    try:
        df = pd.read_excel(input_path)
        base_name = os.path.splitext(filename)[0]
        temp_path = os.path.join(intermediate_subfolder, f"{base_name}_cleaned.xlsx")
        df.to_excel(temp_path, index=False)
    except Exception as e:
        print(f"❌ Failed to clean and convert file {filename}: {e}")
        logging.error(f"Failed to clean and convert file {filename}: {e}")
        return

    wb = load_workbook(temp_path)
    sheet = wb.active

    timestamp_row = None

    for row in sheet.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and "timestamp" in str(cell.value).lower():
                timestamp_row = cell.row
                break
        if timestamp_row:
            break

    if timestamp_row is None:
        print(f"⚠️ Timestamp not found in: {filename}")
        logging.warning(f"Timestamp not found in: {filename}")
        return


    # Infomration Requiring Your Attention (IRYA) section handling
    irya_found = False
    irya_start = 1
    for i in range(1, timestamp_row):
        if sheet.cell(row=i, column=1).value and "Information Requiring Your Attention" in str(sheet.cell(row=i, column=1).value):
            irya_start = i
            irya_found = True
            sheet.cell(row=i, column=1).font = Font(bold=True)

    if irya_found:
        for _ in range(irya_start - 1):
            sheet.delete_rows(1)
        timestamp_row -= (irya_start - 1)
    else:
        for _ in range(timestamp_row - 1):
            sheet.delete_rows(1)
        for _ in range(3):
            sheet.insert_rows(1)
        timestamp_row = 4

    first_data_row = timestamp_row + 1
    last_data_row = sheet.max_row

    for col_idx in reversed(range(2, sheet.max_column + 1)):
        col_values = [sheet.cell(row=r, column=col_idx).value for r in range(first_data_row, last_data_row + 1)]
        if all((v is None or str(v).strip() == "") for v in col_values):
            sheet.delete_cols(col_idx)

    meter_labels = [cell.value for cell in sheet[timestamp_row]]

    for col_idx, cell in enumerate(sheet[timestamp_row], start=1):
        if cell.value:

            original = str(cell.value).strip()
            lines = []

            # Extract leading part (e.g., "06.06ME1")
            match = re.match(r"^([\d.]+[A-Z]*\d*)", original)
            if match:
                number_part = match.group(1)
                lines.append(number_part)
                rest = original[len(number_part):].strip()
            else:
                rest = original

            # Extract unit like (kWh)
            unit = ""
            if "(" in rest and ")" in rest:
                # Clean split and extract
                rest_parts = rest.split("(", 1)
                rest = rest_parts[0].strip()
                unit = f"({rest_parts[1].strip(')')})"

            
            if rest:
                lines.append(rest)
            if unit:
                lines.append(unit)

            formatted = "\n".join(lines)
            cell.value = formatted
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Adjust column width to be tighter
            col_letter = get_column_letter(col_idx)
            max_line_length = max(len(line.strip()) for line in lines)
            sheet.column_dimensions[col_letter].width = max_line_length + 2  # small buffer




    usage_row_index = timestamp_row - 1
    usage_cells = []

    for i, col in enumerate(sheet.iter_cols(min_row=first_data_row, min_col=2), start=2):
        values = [(row, cell.value) for row, cell in enumerate(col, start=first_data_row) if isinstance(cell.value, (int, float))]
        if len(values) < 2:
            continue

        first_row, first = values[0]
        last_row, last = values[-1]
        flip_value = None
        for j in range(1, len(values)):
            prev_val = values[j - 1][1]
            curr_val = values[j][1]
            if prev_val != 0 and abs(curr_val / prev_val) < 0.1:
                flip_value = prev_val
                break

        is_cogen = "cogen" in building_name.lower()
        usage = (round_to_nearest_power_of_10(flip_value, is_cogen) - first + last) if flip_value and ((last < first and not is_cogen) or (last > first and is_cogen)) else last - first

        raw_meter_name = meter_labels[i - 1] if i - 1 < len(meter_labels) else f"Meter {i}"
        clean_meter = extract_clean_meter_name(str(raw_meter_name))

        try:
            first_time = parse(str(sheet.cell(row=first_row, column=1).value))
            last_time = parse(str(sheet.cell(row=last_row, column=1).value))
            next_month = 1 if first_time.month == 12 else first_time.month + 1
            next_year = first_time.year + 1 if first_time.month == 12 else first_time.year
            correct = (
                first_time.day == 1 and first_time.strftime("%I:%M %p") == "12:15 AM" and
                last_time.day == 1 and last_time.month == next_month and last_time.year == next_year and
                last_time.strftime("%I:%M %p") == "12:00 AM"
            )
        except:
            correct = False

        color = "C6EFCE" if correct else "FFC7CE"
        if correct:
            for r in range(first_row + 1, last_row):
                val = sheet.cell(row=r, column=col[0].column).value
                if val is None or str(val).strip() == "":
                    color = "FFFF00"
                    break

        usage_cell = sheet.cell(row=usage_row_index, column=col[0].column)
        usage_cell.value = round(usage, 2)
        usage_cell.number_format = '#,##0.00'
        usage_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        usage_cells.append(usage_cell)
        master_data.append([building_name, clean_meter, round(usage, 2)])


        # Only add summary if it's the "IESO and Hospital" sheet
    if "IESO and Hospital" in filename or "IESO_Hospital" in filename:

        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo

            # Detect last meter data column
            last_meter_col = 1
            for col in range(2, sheet.max_column + 1):
                header_val = sheet.cell(row=timestamp_row, column=col).value
                if header_val and str(header_val).strip():
                    last_meter_col = col



            # Detect how many actual data rows exist under the timestamp
            first_data_row = timestamp_row + 1
            last_data_row = sheet.max_row
            data_row_count = 0
            for r in range(first_data_row, last_data_row + 1):
                has_data = any(
                    sheet.cell(row=r, column=c).value not in [None, ""]
                    for c in range(2, last_meter_col + 1)
                )
                if has_data:
                    data_row_count += 1
                else:
                    break  # Stop at first empty row

                

            row_end = first_data_row + data_row_count - 1

           # print(f"📊 DEBUG – {filename}: table range = A{timestamp_row}:{get_column_letter(last_meter_col)}{row_end}")

            if row_end > timestamp_row and last_meter_col >= 1:
                last_col_letter = get_column_letter(last_meter_col)
                meter_table_ref = f"A{timestamp_row}:{last_col_letter}{row_end}"

                table = Table(displayName=f"MeterTable_{uuid4().hex[:6]}", ref=meter_table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                sheet.add_table(table)
            else:
                print(f"⚠️ Skipping table formatting for {filename}: insufficient data rows.")
        except Exception as e:
            print(f"⚠️ Could not apply table style to IESO file {filename}: {e}")

        ieso_total = 0
        hospital_total = 0

        for cell in usage_cells:
            meter_header = sheet.cell(row=timestamp_row, column=cell.column).value

            
            if not meter_header:
                continue


            normalized = str(meter_header).replace("\n", "").replace(" ", "")

            if "12T1Q1" in normalized or "12T2Q3" in normalized:
                ieso_total += cell.value or 0
            if "12M14A" in normalized or "12M21" in normalized:
                hospital_total += cell.value or 0

        university_total = ieso_total - hospital_total

        base_col = sheet.max_column + 3  # Write 3 columns to the right of existing content
        sheet.cell(row=usage_row_index, column=base_col, value="IESO Purchased:")
        sheet.cell(row=usage_row_index, column=base_col + 1, value=round(ieso_total, 2))

        sheet.cell(row=usage_row_index + 1, column=base_col, value="Hospital Usage:")
        sheet.cell(row=usage_row_index + 1, column=base_col + 1, value=round(hospital_total, 2))

        sheet.cell(row=usage_row_index + 2, column=base_col, value="University Usage:")
        sheet.cell(row=usage_row_index + 2, column=base_col + 1, value=round(university_total, 2))

        # Format the summary values to look clean and readable
        for r in range(usage_row_index, usage_row_index + 3):
            val_cell = sheet.cell(row=r, column=base_col + 1)
            val_cell.number_format = '#,##0.00'
            val_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Widen the summary columns to ensure nothing is squished
        for c in [base_col, base_col + 1]:
            col_letter = get_column_letter(c)
            sheet.column_dimensions[col_letter].width = 20

        # Apply final formatting fix ONLY for IESO and Hospital sheets
    for col in sheet.iter_cols(min_row=usage_row_index + 1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
        sheet.column_dimensions[col_letter].width = max_len + 2  # Ensure values fit cleanly

        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' 
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(cell.value, str) and col[0].column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')  # Timestamp stays left-aligned


    # Apply a table style to the main data if not already styled
    if "IESO and Hospital" not in filename:
        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo

            meter_table_ref = f"A{timestamp_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            table = Table(displayName=f"MeterTable_{uuid4().hex[:6]}", ref=meter_table_ref)

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            sheet.add_table(table)
        except Exception as e:
            print(f"⚠️ Could not apply table style to intermediate file {filename}: {e}")
            logging.error(f"Could not apply table style to intermediate file {filename}: {e}")

    
    

    # Dynamically determine bill_month from first datetime under timestamp
    bill_month_final = bill_month  # default fallback
    for row in range(timestamp_row + 1, sheet.max_row + 1):
        val = sheet.cell(row=row, column=1).value
        try:
            dt = parse(str(val), fuzzy=True)
            bill_month_final = dt.strftime("%B")
            break
        except:
            continue

    output_filename = f"{today_str}_{time_str}_{bill_month_final}_{building_name}.xlsx"

    output_path = os.path.join(intermediate_subfolder, output_filename)
    wb.save(output_path)
    if os.path.exists(temp_path) and temp_path != input_path:
        os.remove(temp_path)
    print(f"✅ Completed file: {filename}")
    logging.info(f"Completed file: {filename} - saved to {output_path}")

def main():
    now = datetime.now()
    master_data = []
    today_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H-%M")
    
    bill_month = now.strftime("%B")
    master_data = []

    # Use TEMP name for now — will rename after processing files
    intermediate_subfolder_name = f"{today_str}_{time_str}_TEMP"
    intermediate_subfolder = os.path.join(INTERMEDIATE_FOLDER, intermediate_subfolder_name)
    os.makedirs(intermediate_subfolder, exist_ok=True)

    files = [file for file in os.listdir(PRE_UPDATED) if file.endswith((".xls", ".xlsx"))]
    print(f"📁 Found {len(files)} files in Pre-Updated folder.")
    logging.info(f"Found {len(files)} files in Pre-Updated folder.")


    for idx, file in enumerate(files, start=1):
        building = clean_building_name(file)
        file_path = os.path.join(PRE_UPDATED, file)
        
        if re.match(r"^nitrogen\s+\d+[a-z]?$", os.path.splitext(file)[0].lower()):
            handle_nitrogen_file(file_path, intermediate_subfolder, building, today_str, bill_month, time_str, master_data)
        else:
            format_excel(file_path, intermediate_subfolder, master_data, building, today_str, bill_month, time_str)
        
        os.remove(file_path)
        print(f"📦 {idx}/{len(files)} files processed.")
        logging.info(f"Processed file {idx}/{len(files)}: {file}")


    # Determine most frequent billing month from intermediate file names
    from collections import Counter #library imported
    month_counts = Counter()
    for file in os.listdir(intermediate_subfolder):
        match = re.search(rf"{time_str}_(\w+)_", file)
        if match:
            month_counts[match.group(1)] += 1

    if month_counts:
        most_common_month = month_counts.most_common(1)[0][0]
    else:
        most_common_month = bill_month  # if this fails then default to current month 

    # Rename intermediate folder to include most common billing month
    final_folder_name = f"{today_str}_{time_str}_{most_common_month}"
    final_folder_path = os.path.join(INTERMEDIATE_FOLDER, final_folder_name)
    os.rename(intermediate_subfolder, final_folder_path)
    intermediate_subfolder = final_folder_path  # update reference for later use


    if master_data:
        df = pd.DataFrame(master_data, columns=["Building", "Meter", "Usage"])
        df["Usage"] = df["Usage"].map(lambda x: f"{x:,.2f}")
        grouped = df.groupby("Building")
        final_rows = []
        for name, group in grouped:
            final_rows.append([None, None, None])  # Spacer row
            final_rows.extend(group.values.tolist())
            final_rows.append([None, None, None])  # Spacer row

        styled_df = pd.DataFrame(final_rows, columns=["Building", "Meter", "Usage"])
        master_filename = f"Final-{today_str}-{time_str}-{most_common_month}.xlsx"
        master_path = os.path.join(OUTPUT_FOLDER, master_filename)



        with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
            styled_df.to_excel(writer, index=False)
            sheet = writer.sheets['Sheet1']



            color1 = "DAECF9"
            color2 = "B9D9F7"
            current_fill = color1
            start = 2  # Skip header

            while start <= sheet.max_row:
                building_name = sheet.cell(row=start, column=1).value
                if not building_name:
                    start += 1
                    continue

                # Apply fill color block
                end = start
                while end <= sheet.max_row and sheet.cell(row=end, column=1).value:
                    end += 1
                for r in range(start, end):
                    for c in range(1, 4):
                        sheet.cell(row=r, column=c).fill = PatternFill(start_color=current_fill, end_color=current_fill, fill_type="solid")
                current_fill = color2 if current_fill == color1 else color1
                start = end + 1

            
            # Adjust column widths and align Usage column to the right
            for column_cells in sheet.columns:
                col_letter = get_column_letter(column_cells[0].column)
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                sheet.column_dimensions[col_letter].width = max_length + 1

                for cell in column_cells:
                    if column_cells[0].column == 3:  # Column C = "Usage"
                        cell.alignment = Alignment(horizontal="right", vertical="center")


        print(f"✅ Master Excel file saved: {master_filename}")
        logging.info(f"Master Excel file saved: {master_filename} - {master_path}")

if __name__ == "__main__":
    main()
    input("Press Enter to exit...")

